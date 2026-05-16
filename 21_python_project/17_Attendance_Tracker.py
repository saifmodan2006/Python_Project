"""Attendance tracker with safer behavior and test mode.

Usage examples:
  # Create a small sample workbook for testing
  python 17_Attendance_Tracker.py --create-sample

  # Mark students 1 and 3 absent for CI (subject 'ci') in test mode (no emails)
  python 17_Attendance_Tracker.py --workbook attendance.xlsx --subject ci --absentees "1 3" --test

Notes:
 - This script defaults to test mode (no emails). To enable sending emails, provide SMTP settings and --send-emails (not recommended to store passwords in code).
"""

import argparse
import os
from pathlib import Path
from typing import List, Optional

import openpyxl


SUBJECT_COLUMNS = {
    "ci": 3,
    "python": 4,
    "dm": 5,
}


def create_sample_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Header: Roll, Email, CI, Python, DM
    ws.append(["Roll", "Email", "CI", "Python", "DM"])
    sample = [
        (1, "student1@example.com", 0, 0, 0),
        (2, "student2@example.com", 0, 0, 0),
        (3, "student3@example.com", 0, 0, 0),
    ]
    for row in sample:
        ws.append(list(row))
    wb.save(path)
    print(f"Sample workbook created at {path}")


def load_workbook(path: Path):
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    return wb, ws


def find_student_row(ws, roll: int) -> Optional[int]:
    # assume roll numbers are in column 1 starting at row 2
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        try:
            if int(val) == roll:
                return r
        except Exception:
            continue
    return None


def mark_absentees(ws, absentees: List[int], subject: str, threshold: int = 2) -> dict:
    """Mark absentees in worksheet. Returns summary dict with actions.

    Summary includes lists of students at threshold and above threshold.
    """
    col = SUBJECT_COLUMNS.get(subject.lower())
    if col is None:
        raise ValueError(f"Unknown subject: {subject}")

    at_threshold = []
    above_threshold = []

    for roll in absentees:
        row = find_student_row(ws, roll)
        if row is None:
            print(f"Warning: roll {roll} not found in workbook")
            continue
        cell = ws.cell(row=row, column=col)
        prev = cell.value or 0
        new = prev + 1
        cell.value = new
        if new == threshold:
            at_threshold.append((roll, ws.cell(row=row, column=2).value))
        elif new > threshold:
            above_threshold.append((roll, ws.cell(row=row, column=2).value))

    return {"at_threshold": at_threshold, "above_threshold": above_threshold}


def send_notifications(summary: dict, subject: str, test: bool = True, smtp_config: Optional[dict] = None) -> None:
    """In test mode, print what would be sent. If test=False and smtp_config provided, would send emails.
    This implementation intentionally does not send emails by default to protect credentials.
    """
    title = f"Attendance alert for {subject}"
    if summary["at_threshold"]:
        print("Students reached threshold:")
        for roll, email in summary["at_threshold"]:
            msg = f"Student {roll} ({email}) has reached the absence threshold for {subject}. Please warn them."
            if test:
                print("[TEST]", msg)
            else:
                print("Would send email to student:", email)
                # implement actual sending here if desired

    if summary["above_threshold"]:
        print("Students above threshold (notify staff):")
        rolls = ", ".join(str(r) for r, _ in summary["above_threshold"])
        print(f"Students: {rolls}")
        # In test mode we just print staff notification
        if test:
            print(f"[TEST] Notify staff: {len(summary['above_threshold'])} students above threshold for {subject}")
        else:
            print("Would send staff email with list:", rolls)


def parse_args():
    parser = argparse.ArgumentParser(description="Attendance tracker (safe/test mode)")
    parser.add_argument("--workbook", type=str, default="attendance.xlsx", help="Path to Excel workbook")
    parser.add_argument("--create-sample", action="store_true", help="Create a sample workbook and exit")
    parser.add_argument("--subject", type=str, choices=list(SUBJECT_COLUMNS.keys()), help="Subject to mark (ci, python, dm)")
    parser.add_argument("--absentees", type=str, help="Space-separated roll numbers to mark absent, e.g. '1 3 5'")
    parser.add_argument("--threshold", type=int, default=2, help="Absence threshold")
    parser.add_argument("--save", action="store_true", help="Save workbook after changes")
    parser.add_argument("--test", action="store_true", help="Test mode: do not send emails; print actions instead")
    return parser.parse_args()


def main():
    args = parse_args()
    wb_path = Path(args.workbook)

    if args.create_sample:
        create_sample_workbook(wb_path)
        return

    if not args.subject or not args.absentees:
        print("Error: --subject and --absentees are required unless --create-sample is used")
        return

    # parse absentees
    absentees = []
    try:
        absentees = [int(x) for x in args.absentees.split()]
    except Exception:
        print("Error parsing --absentees. Provide space-separated integers like '1 3 5'")
        return

    try:
        wb, ws = load_workbook(wb_path)
    except Exception as e:
        print("Error loading workbook:", e)
        return

    summary = mark_absentees(ws, absentees, args.subject, threshold=args.threshold)

    send_notifications(summary, args.subject, test=args.test)

    if args.save:
        wb.save(wb_path)
        print(f"Workbook saved to {wb_path}")


if __name__ == '__main__':
    main()
